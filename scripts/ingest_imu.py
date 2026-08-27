"""Ingest the SWEAT-DTG pelvis IMU export into a tidy, tagged dataset and
locate each trial's start/end within the recording.

Each of the 149 sheets in ``SWEAT_DTG IMU Summary Data_ALL.xlsx`` holds one
participant x condition trial (e.g. ``SWEAT_001DTG1``): a block of ``//``
metadata lines followed by a header row (``PacketCounter, [SampleTimeFine,]
Acc_X, Acc_Y, Acc_Z, Roll, Pitch, Yaw``) and ~100 Hz samples.

There is no independent marker for gait start/end in the recording, and many
sheets contain substantially more than the timed 3 m walk (extra idle time,
repeated attempts) with no reliable quiet gap separating it from the real
trial — so a pure quiet-stance/amplitude threshold (mean + k*SD on rolling
Acc-magnitude SD, the "2SD change" rule) alone often grabs the wrong window
(see `detect_quiet_periods` / the `bout_*` columns in the summary).

Instead the primary method anchors on the mid-trial turn: the DTG protocol
is a 3 m walk-with-turn, so the turn should sit at roughly the midpoint of
the timed walk. The turn is a large, brief, easy-to-find spike in
pseudo-angular-rate (`gyro_mag`, differentiated from the orientation
channels since no raw gyro is exported); the trial window is then
turn_time +/- logged_duration/2, with everything outside it discarded as
excess (`detect_trial_window`). This falls back to the amplitude-bout window
when a trial has no logged duration to anchor against.

Outputs (under --outdir, default "output/"):
  - imu_samples_tagged.parquet   long-format, every sample from every trial,
                                  tagged with participant/condition and a
                                  phase label (excess_before / trial /
                                  excess_after).
  - trial_summary.csv            one row per trial: turn-anchored window,
                                  the raw amplitude-bout window for
                                  comparison, logged duration, and QC flags
                                  (window_clipped, n_bouts, etc).
  - diagnostic_plots/*.png       a handful of example trials with both
                                  methods and the detected turn overlaid.

Usage:
    python scripts/ingest_imu.py
    python scripts/ingest_imu.py --k 2.5 --plot-n 12
"""
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, asdict
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

FS_DEFAULT = 100.0  # Hz. Confirmed from continuous PacketCounter increments
                     # against known trial durations (Trial Times spreadsheet).

SHEET_RE = re.compile(r"^SWEAT_(\d{3})\s*(DTG\s*TRIAL|DTG[123])$", re.IGNORECASE)


def parse_sheet_name(sheet_name: str) -> tuple[str, str]:
    """'SWEAT_001DTG1' -> ('SWEAT-001', 'DTG1'); 'SWEAT_005DTG TRIAL' -> ('SWEAT-005', 'DTG_TRIAL')."""
    m = SHEET_RE.match(sheet_name.strip())
    if not m:
        raise ValueError(f"Unrecognized sheet name format: {sheet_name!r}")
    pid, cond = m.groups()
    cond = re.sub(r"\s+", "", cond).upper()  # 'DTG TRIAL' -> 'DTGTRIAL'
    if cond == "DTGTRIAL":
        cond = "DTG_TRIAL"
    return f"SWEAT-{pid}", cond


def load_trial_raw(ws) -> pd.DataFrame:
    """Read one worksheet, skip the '//' metadata block, return raw columns."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "PacketCounter")
    header = [h for h in rows[hdr_idx] if h is not None]
    data = rows[hdr_idx + 1:]
    df = pd.DataFrame(data, columns=list(rows[hdr_idx][: len(header)]))
    df = df[header]  # drop any trailing all-None columns
    if "SampleTimeFine" in df.columns:
        df = df.drop(columns=["SampleTimeFine"])  # unused/empty in this export
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def add_derived_signals(df: pd.DataFrame, fs: float) -> pd.DataFrame:
    df = df.copy()
    n = len(df)
    df["t_s"] = np.arange(n) / fs
    df["acc_mag"] = np.sqrt(df["Acc_X"] ** 2 + df["Acc_Y"] ** 2 + df["Acc_Z"] ** 2)

    # Pseudo-angular-rate from the orientation channels (no raw gyro is
    # exported). Unwrap first so genuine continuous rotation isn't mistaken
    # for a jump at the +-180 deg boundary. NOTE: Roll/Pitch/Yaw are Euler
    # angles and can still swing rapidly near gimbal-lock (pitch ~ +-90 deg,
    # observed around the mid-trial turn) — treat this signal as informative
    # for the gait/turn portion, not a robust quiet-detector by itself, hence
    # detect_quiet() below is driven by acc_mag rather than this.
    roll_u = np.degrees(np.unwrap(np.radians(df["Roll"].to_numpy())))
    pitch_u = np.degrees(np.unwrap(np.radians(df["Pitch"].to_numpy())))
    yaw_u = np.degrees(np.unwrap(np.radians(df["Yaw"].to_numpy())))
    t = df["t_s"].to_numpy()
    df["roll_rate"] = np.gradient(roll_u, t) if n > 1 else 0.0
    df["pitch_rate"] = np.gradient(pitch_u, t) if n > 1 else 0.0
    df["yaw_rate"] = np.gradient(yaw_u, t) if n > 1 else 0.0
    df["gyro_mag"] = np.sqrt(df["roll_rate"] ** 2 + df["pitch_rate"] ** 2 + df["yaw_rate"] ** 2)
    return df


def rolling_std(x: np.ndarray, win: int) -> np.ndarray:
    win = max(win, 1)
    s = pd.Series(x).rolling(window=win, center=True, min_periods=1).std(ddof=0)
    return s.to_numpy()


@dataclass
class Bout:
    start_idx: int
    end_idx: int  # inclusive
    duration_s: float


@dataclass
class QuietDetectionResult:
    onset_idx: int
    offset_idx: int
    start_detected: bool
    end_detected: bool
    offset_source: str  # 'detected' | 'end_of_recording'
    baseline_mean: float
    baseline_sd: float
    threshold: float
    n_bouts: int
    other_bouts: list  # Bout objects for any extra movement bouts, e.g. repeated reps
    bouts: list  # every bout found (primary + other), for building an activity mask


@dataclass
class TrialWindowResult:
    onset_idx: int
    offset_idx: int
    turn_idx: int | None
    turn_peak_gyro_mag: float | None
    method: str  # 'turn_anchored' | 'bout_fallback'
    clipped: bool  # half-duration window ran off the start/end of the recording


def _global_quiet_baseline(acc_sd: np.ndarray, low_percentile: float = 20.0) -> tuple[float, float]:
    """Estimate the sensor's stationary noise floor for this trial from the
    lower tail of the rolling-SD distribution across the WHOLE recording,
    rather than a single scanned window — a single window's own internal
    variance is a fragile (noisy) sd estimate and can blow the threshold up
    if that window happens to catch one small transient. Quiet stance
    (start, end, or an idle stretch mid-recording) reliably makes up the
    bottom of the distribution in every trial inspected."""
    cutoff = np.percentile(acc_sd, low_percentile)
    low_vals = acc_sd[acc_sd <= cutoff]
    return float(low_vals.mean()), float(low_vals.std())


def _find_bouts(active: np.ndarray, merge_gap_n: int, min_bout_n: int) -> list[Bout]:
    """Turn a boolean 'is this sample active' array into a list of bouts:
    bridge short quiet gaps inside a bout (merge_gap_n), then drop bouts
    shorter than min_bout_n (noise blips / knocks, not real gait)."""
    n = len(active)
    idx = np.flatnonzero(active)
    if len(idx) == 0:
        return []
    # merge runs separated by small gaps
    gaps = np.flatnonzero(np.diff(idx) > merge_gap_n)
    starts = np.concatenate(([idx[0]], idx[gaps + 1]))
    ends = np.concatenate((idx[gaps], [idx[-1]]))
    bouts = [Bout(int(s), int(e), 0.0) for s, e in zip(starts, ends) if (e - s + 1) >= min_bout_n]
    return bouts


def detect_quiet_periods(
    df: pd.DataFrame,
    fs: float,
    k: float = 2.0,
    roll_window_s: float = 0.25,
    merge_gap_s: float = 0.5,
    min_bout_s: float = 1.0,
) -> QuietDetectionResult:
    """Segment the trial into quiet/active bouts from Acc magnitude, using a
    single threshold = baseline_mean + k*baseline_sd (the "2SD" rule) taken
    from the lower tail of the rolling-SD distribution across the whole
    recording — not just the first/last couple of seconds, since some sheets
    run on well past the end of the timed trial (idle IMU) with occasional
    movement blips in that tail that a start/end-only search mistakes for
    the real trial boundary.

    If more than one bout survives (e.g. a repeated attempt, per the study
    notes' "repeated rep" case), the LONGEST bout is picked as primary — the
    real trial (plus whatever extra idle/movement time the recording bundles
    in with it) is reliably the single longest continuous non-quiet stretch,
    whereas repeated-attempt reps and incidental blips are short. Picking by
    closest-absolute-match to the logged duration instead was tried and
    rejected: a short incidental blip can be numerically closer to the
    logged duration than the real (but padded) trial bout, which then feeds
    a wrong search window to detect_trial_window's turn search. The rest of
    the bouts are kept as `other_bouts` for QC rather than silently
    discarded.
    """
    n = len(df)
    roll_n = max(int(round(roll_window_s * fs)), 1)
    merge_gap_n = max(int(round(merge_gap_s * fs)), 1)
    min_bout_n = max(int(round(min_bout_s * fs)), 1)

    acc_sd = rolling_std(df["acc_mag"].to_numpy(), roll_n)
    baseline_mean, baseline_sd = _global_quiet_baseline(acc_sd)
    threshold = baseline_mean + k * baseline_sd

    bouts = _find_bouts(acc_sd > threshold, merge_gap_n, min_bout_n)
    for b in bouts:
        b.duration_s = (b.end_idx - b.start_idx + 1) / fs

    if not bouts:
        # no activity distinguishable from noise floor at all
        return QuietDetectionResult(0, n - 1, False, False, "end_of_recording",
                                     baseline_mean, baseline_sd, threshold, 0, [], [])

    primary = max(bouts, key=lambda b: b.duration_s)
    other = [b for b in bouts if b is not primary]

    start_detected = primary.start_idx > 0
    end_detected = primary.end_idx < n - 1
    offset_source = "detected" if end_detected else "end_of_recording"

    return QuietDetectionResult(
        onset_idx=primary.start_idx,
        offset_idx=primary.end_idx,
        start_detected=start_detected,
        end_detected=end_detected,
        offset_source=offset_source,
        baseline_mean=baseline_mean,
        baseline_sd=baseline_sd,
        threshold=threshold,
        n_bouts=len(bouts),
        other_bouts=other,
        bouts=bouts,
    )


def find_turn_idx(gyro_mag: np.ndarray, active_mask: np.ndarray | None) -> tuple[int, float]:
    """Locate the mid-trial turn as the point of peak pseudo-angular-rate
    magnitude. Restricted to samples inside a detected movement bout so an
    idle stretch (or a completely unrelated blip) can't be mistaken for it —
    the turn is a real, large, brief reorientation, and only makes sense
    while the participant is already walking."""
    if active_mask is not None and active_mask.any():
        candidate = np.where(active_mask, gyro_mag, -np.inf)
    else:
        candidate = gyro_mag
    idx = int(np.argmax(candidate))
    return idx, float(gyro_mag[idx])


def detect_trial_window(
    df: pd.DataFrame,
    fs: float,
    bout_result: QuietDetectionResult,
    logged_duration_s: float | None,
) -> TrialWindowResult:
    """Anchor the trial window on the mid-trial turn rather than on
    quiet-period boundaries: the DTG protocol is a 3 m walk-with-turn, so the
    turn should sit at roughly the midpoint of the timed trial. Recordings
    often contain substantially more than the timed walk (extra idle time,
    repeated attempts) with no reliable quiet gap separating it from the
    real trial, which defeats amplitude-threshold segmentation alone —
    but the turn itself is a distinctive, brief event that survives that
    problem. Take turn_time +/- logged_duration/2 as the trial window and
    treat everything outside it as excess data to discard.

    Falls back to the amplitude-bout result when no logged duration is
    available to anchor against.
    """
    n = len(df)
    if logged_duration_s is None or np.isnan(logged_duration_s) or not bout_result.bouts:
        return TrialWindowResult(bout_result.onset_idx, bout_result.offset_idx, None, None,
                                  "bout_fallback", False)

    # Restrict the turn search to the amplitude method's own best-matching
    # bout (bout_result.onset_idx/offset_idx), not the union of every bout:
    # a later, unrelated blip (equipment adjustment, repeated attempt) can
    # have a larger pseudo-gyro spike than the real turn simply because
    # Euler-angle differentiation amplifies unpredictably near gimbal lock,
    # so searching every bout equally lets it hijack the turn estimate.
    # Also trim a margin off each end of that bout: when the bout is padded
    # with extra time, the largest spike right at its edge is usually the
    # participant starting/stopping/removing the sensor, not the turn — a
    # genuine mid-walk turn is, by definition, not at the very boundary of
    # continuous movement.
    bout_len = bout_result.offset_idx - bout_result.onset_idx + 1
    margin_n = min(int(round(max(1.5, 0.1 * bout_len / fs) * fs)), bout_len // 2 - 1) if bout_len > 2 else 0
    margin_n = max(margin_n, 0)
    mask = np.zeros(n, dtype=bool)
    mask[bout_result.onset_idx + margin_n: bout_result.offset_idx + 1 - margin_n] = True
    turn_idx, turn_peak = find_turn_idx(df["gyro_mag"].to_numpy(), mask)

    half_n = int(round(logged_duration_s / 2 * fs))
    raw_onset = turn_idx - half_n
    raw_offset = turn_idx + half_n
    onset = max(raw_onset, 0)
    offset = min(raw_offset, n - 1)
    clipped = (onset != raw_onset) or (offset != raw_offset)

    return TrialWindowResult(onset, offset, turn_idx, turn_peak, "turn_anchored", clipped)


def tag_phases(df: pd.DataFrame, window: TrialWindowResult) -> pd.Series:
    """Tag every sample relative to the trial window: 'trial' is what
    downstream gait-metric code should use; 'excess_before'/'excess_after'
    is everything discarded as outside the turn-anchored +/- half-duration
    window (setup, repeated attempts, idle IMU, etc.)."""
    n = len(df)
    phase = np.full(n, "excess_before", dtype=object)
    phase[window.onset_idx: window.offset_idx + 1] = "trial"
    phase[window.offset_idx + 1:] = "excess_after"
    return pd.Series(phase, index=df.index, name="phase")


def load_trial_times(path: Path) -> pd.DataFrame:
    """Melt the Trial Times spreadsheet to (participant, condition, logged_duration_s)."""
    raw = pd.read_excel(path, sheet_name="DTG Trial Times", header=4)
    raw = raw.rename(columns={raw.columns[0]: "Date", raw.columns[1]: "Participant",
                               raw.columns[2]: "DTG_TRIAL", raw.columns[3]: "DTG1",
                               raw.columns[4]: "DTG2", raw.columns[5]: "DTG3",
                               raw.columns[6]: "Notes"})
    raw = raw[raw["Participant"].astype(str).str.startswith("SWEAT")]
    long = raw.melt(id_vars=["Participant", "Notes"], value_vars=["DTG_TRIAL", "DTG1", "DTG2", "DTG3"],
                     var_name="condition", value_name="logged_duration_s")
    long = long.rename(columns={"Participant": "participant"})
    return long[["participant", "condition", "logged_duration_s", "Notes"]]


def process_all(source_xlsx: Path, trial_times_xlsx: Path, outdir: Path, k: float,
                 fs: float, plot_n: int) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    plot_dir = outdir / "diagnostic_plots"
    plot_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(source_xlsx, read_only=True, data_only=True)
    trial_times = load_trial_times(trial_times_xlsx)

    tagged_frames = []
    summary_rows = []
    plotted = 0

    for sheet_name in wb.sheetnames:
        try:
            participant, condition = parse_sheet_name(sheet_name)
        except ValueError as e:
            print(f"skip {sheet_name!r}: {e}")
            continue

        raw = load_trial_raw(wb[sheet_name])
        if raw.empty:
            print(f"skip {sheet_name!r}: no data rows")
            continue

        df = add_derived_signals(raw, fs)

        logged = trial_times[(trial_times.participant == participant) & (trial_times.condition == condition)]
        logged_duration_s = float(logged["logged_duration_s"].iloc[0]) if len(logged) and pd.notna(logged["logged_duration_s"].iloc[0]) else np.nan
        notes = logged["Notes"].iloc[0] if len(logged) else None

        bout_result = detect_quiet_periods(df, fs, k=k)
        window = detect_trial_window(df, fs, bout_result, logged_duration_s)
        df["phase"] = tag_phases(df, window)
        df.insert(0, "condition", condition)
        df.insert(0, "participant", participant)
        df.insert(0, "sheet", sheet_name)
        tagged_frames.append(df)

        detected_duration_s = (window.offset_idx - window.onset_idx) / fs
        bout_duration_s = (bout_result.offset_idx - bout_result.onset_idx) / fs
        summary_rows.append({
            "sheet": sheet_name,
            "participant": participant,
            "condition": condition,
            "n_samples": len(df),
            "fs_hz": fs,
            "logged_duration_s": logged_duration_s,
            "window_method": window.method,
            "turn_time_s": window.turn_idx / fs if window.turn_idx is not None else np.nan,
            "turn_peak_gyro_mag": window.turn_peak_gyro_mag,
            "onset_s": window.onset_idx / fs,
            "offset_s": window.offset_idx / fs,
            "detected_duration_s": detected_duration_s,
            "duration_diff_s": detected_duration_s - logged_duration_s if pd.notna(logged_duration_s) else np.nan,
            "window_clipped": window.clipped,
            "bout_onset_s": bout_result.onset_idx / fs,
            "bout_offset_s": bout_result.offset_idx / fs,
            "bout_duration_s": bout_duration_s,
            "bout_duration_diff_s": bout_duration_s - logged_duration_s if pd.notna(logged_duration_s) else np.nan,
            "n_bouts": bout_result.n_bouts,
            "trial_notes": notes,
        })

        if plot_n and plotted < plot_n:
            plot_trial(df, window, bout_result, sheet_name, fs, plot_dir)
            plotted += 1

        turn_time_s = window.turn_idx / fs if window.turn_idx is not None else float("nan")
        print(f"{sheet_name:22s} n={len(df):5d} ({window.method}) onset={window.onset_idx/fs:5.2f}s "
              f"offset={window.offset_idx/fs:5.2f}s turn={turn_time_s:5.2f}s "
              f"dur={detected_duration_s:5.2f}s logged={logged_duration_s}"
              f"{' [CLIPPED]' if window.clipped else ''}")

    tagged = pd.concat(tagged_frames, ignore_index=True)
    tagged.to_parquet(outdir / "imu_samples_tagged.parquet", index=False)

    summary = pd.DataFrame(summary_rows)
    summary.to_csv(outdir / "trial_summary.csv", index=False)

    n_turn_anchored = (summary["window_method"] == "turn_anchored").sum()
    n_clipped = summary["window_clipped"].sum()
    within_tol = (summary["duration_diff_s"].abs() <= 0.5).sum()
    print()
    print(f"Wrote {len(tagged):,} tagged samples across {len(summary)} trials to {outdir}")
    print(f"Turn-anchored window used for {n_turn_anchored}/{len(summary)} trials "
          f"({len(summary) - n_turn_anchored} fell back to the amplitude-bout window, no logged duration available)")
    print(f"{n_clipped} trials had the turn so close to the recording edge that the "
          f"+/-half-duration window had to be clipped — review these (window_clipped=True in trial_summary.csv)")
    print(f"{within_tol}/{len(summary)} trials landed within 0.5s of the logged duration by construction; "
          f"see bout_duration_diff_s for how far the raw amplitude-bout estimate was off before turn-anchoring")


def plot_trial(df: pd.DataFrame, window: TrialWindowResult, bout_result: QuietDetectionResult,
               sheet_name: str, fs: float, plot_dir: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(3, 1, figsize=(11, 8), sharex=True)
    t = df["t_s"].to_numpy()
    axes[0].plot(t, df["acc_mag"])
    axes[0].set_ylabel("|Acc| (m/s^2)")
    axes[0].set_title(sheet_name)

    roll_n = max(int(round(0.25 * fs)), 1)
    axes[1].plot(t, rolling_std(df["acc_mag"].to_numpy(), roll_n))
    axes[1].set_ylabel("rolling SD |Acc|")

    axes[2].plot(t, df["gyro_mag"], color="tab:purple")
    axes[2].set_ylabel("pseudo-gyro mag (deg/s)")
    axes[2].set_xlabel("time (s)")

    for ax in axes:
        ax.axvline(window.onset_idx / fs, color="green", linestyle="--", label="turn-anchored onset")
        ax.axvline(window.offset_idx / fs, color="green", linestyle="--", label="turn-anchored offset")
        ax.axvline(bout_result.onset_idx / fs, color="orange", linestyle=":", label="amplitude-bout onset")
        ax.axvline(bout_result.offset_idx / fs, color="orange", linestyle=":", label="amplitude-bout offset")
        if window.turn_idx is not None:
            ax.axvline(window.turn_idx / fs, color="red", linestyle="-", linewidth=1, label="turn")
        ax.grid(alpha=0.3)
    axes[1].axhline(bout_result.threshold, color="gray", linestyle=":", linewidth=1, label="amplitude threshold")
    axes[0].legend(loc="upper right", fontsize=7)
    plt.tight_layout()
    plt.savefig(plot_dir / f"{sheet_name}.png", dpi=110)
    plt.close(fig)


def main():
    here = Path(__file__).resolve().parent.parent
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--source", type=Path, default=here / "SWEAT_DTG IMU Summary Data_ALL.xlsx")
    p.add_argument("--trial-times", type=Path, default=here / "SWEAT_DTG Trial times_ALL.xlsx")
    p.add_argument("--outdir", type=Path, default=here / "output")
    p.add_argument("--fs", type=float, default=FS_DEFAULT, help="nominal sample rate (Hz)")
    p.add_argument("--k", type=float, default=2.0, help="baseline_mean + k*baseline_sd threshold multiplier")
    p.add_argument("--plot-n", type=int, default=10, help="number of diagnostic plots to save (0 to disable)")
    args = p.parse_args()

    process_all(args.source, args.trial_times, args.outdir, k=args.k, fs=args.fs, plot_n=args.plot_n)


if __name__ == "__main__":
    main()
